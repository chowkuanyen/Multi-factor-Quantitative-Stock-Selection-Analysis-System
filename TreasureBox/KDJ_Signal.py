import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
import warnings
from ConfigParser import Config  # 确保 Config.py 在同一环境或路径下

if __name__ == "__main__":
    # --- 1. 配置与路径初始化 ---
    # 获取当前脚本目录 (TreasureBox)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # 获取项目根目录 (BAISYS_QUANT)
    root_dir = os.path.dirname(current_dir)
    # 构建配置文件路径
    config_path = os.path.join(root_dir, 'config.ini')

    # 加载配置
    try:
        config = Config(config_path)
        print(f"✅ 成功加载配置文件: {config_path}")
    except Exception as e:
        print(f"❌ 配置加载失败: {e}")
        sys.exit(1)

    # --- 2. 数据库连接构建 ---
    DB_URI = f"postgresql://{config.DB_USER}:{config.DB_PASSWORD}@{config.DB_HOST}:{config.DB_PORT}/{config.DB_NAME}"
    print(f"📊 数据库连接准备就绪")

    # --- 3. 报告路径设置 ---
    # 读取配置文件中的 home_directory 并展开 ~
    REPORT_OUTPUT_DIR = os.path.expanduser(config.HOME_DIRECTORY)
    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)  # 确保路径存在
    print(f"📁 报告将保存至: {REPORT_OUTPUT_DIR}")

    # --- 4. 业务逻辑：数据计算 ---
    # 计算周期 (15天前)
    one_month_ago = (datetime.today() - timedelta(days=15)).strftime('%Y-%m-%d')
    print(f"📅 统计周期: {one_month_ago} 至今天")

    # 创建引擎
    engine = create_engine(DB_URI)

    # --- 4.1 查询有信号的股票 ---
    query_signal_stocks = f"""
    SELECT DISTINCT stock_code, stock_name 
    FROM app_stock_strategy_report 
    WHERE kdj_signal IS NOT NULL 
    AND archive_date >= '{one_month_ago}' 
    ORDER BY stock_code;
    """
    df_signal_stocks = pd.read_sql(query_signal_stocks, engine)
    engine.dispose()

    if df_signal_stocks.empty:
        print("⚠️ 未找到有非空 kdj_signal 的股票。")
        exit()


    # --- 4.2 股票代码映射处理 (添加交易所前缀) ---
    def add_exchange_prefix(stock_code):
        stock_str = str(stock_code).strip()
        if stock_str.startswith('6'):
            return f"sh{stock_str}"
        elif stock_str.startswith('0'):
            return f"sz{stock_str}"
        elif stock_str.startswith('8'):
            return f"bj{stock_str}"
        else:
            return f"sz{stock_str}"  # 默认归为深市


    # 构建映射
    stock_info_map = {}
    for _, row in df_signal_stocks.iterrows():
        code = row['stock_code']
        name = row['stock_name'] if row['stock_name'] else "未知名称"
        prefixed_symbol = add_exchange_prefix(code)
        stock_info_map[code] = {
            'name': name,
            'symbol': prefixed_symbol
        }

    # --- 4.3 筛选有价格数据的股票 ---
    signal_stock_codes = list(stock_info_map.keys())
    print(f"🔍 从策略报告中找到 {len(signal_stock_codes)} 只有信号的股票")
    print("正在从 stock_daily_kline 查询这些股票的近30天价格数据...")

    engine = create_engine(DB_URI)
    prefixed_symbols = [stock_info_map[code]['symbol'] for code in signal_stock_codes]
    in_clause = "', '".join(prefixed_symbols)

    query_kline = f"""
    SELECT DISTINCT symbol 
    FROM stock_daily_kline 
    WHERE symbol IN ('{in_clause}') 
    AND trade_date >= '{one_month_ago}' 
    AND "close" IS NOT NULL;
    """

    df_valid_symbols = pd.read_sql(query_kline, engine)
    engine.dispose()

    valid_symbols_set = set(df_valid_symbols['symbol'].tolist())
    print(f"✅ 在 stock_daily_kline 中找到 {len(valid_symbols_set)} 只有实际价格数据的股票")

    # 构建有效股票池
    effective_stock_codes = []
    for code in signal_stock_codes:
        symbol = stock_info_map[code]['symbol']
        if symbol in valid_symbols_set:
            effective_stock_codes.append(code)

    print(f"✅ 初步有效股票数（有信号 + 有价格）：{len(effective_stock_codes)}")
    print(f"🧹 已剔除 {len(signal_stock_codes) - len(effective_stock_codes)} 个无价格数据的无效股票")

    if len(effective_stock_codes) == 0:
        print("⚠️ 没有股票同时满足‘有KDJ信号’且‘有价格数据’。")
        exit()

    # --- 4.4 获取信号日及涨幅计算 ---
    print("⏳ 正在查询每只股票的最后一次KDJ信号日期...")
    engine = create_engine(DB_URI)
    query_last_signal_date = f"""
    SELECT stock_code, MAX(archive_date) AS last_signal_date 
    FROM app_stock_strategy_report 
    WHERE kdj_signal IS NOT NULL 
    AND archive_date >= '{one_month_ago}' 
    GROUP BY stock_code 
    ORDER BY stock_code;
    """
    df_last_signal_date = pd.read_sql(query_last_signal_date, engine)
    engine.dispose()

    # 映射信号日期
    last_signal_date_map = {}
    for _, row in df_last_signal_date.iterrows():
        code = row['stock_code']
        last_signal_date_map[code] = row['last_signal_date'].strftime('%Y-%m-%d')

    print("💰 正在获取每只股票在信号日的收盘价...")
    signal_dates_list = []
    for code in effective_stock_codes:
        symbol = stock_info_map[code]['symbol']
        date_str = last_signal_date_map[code]
        signal_dates_list.append((code, symbol, date_str))

    # 查询信号日收盘价
    if not signal_dates_list:
        print("⚠️ 无有效信号日数据。")
        exit()

    symbol_date_pairs = []
    for code, symbol, date_str in signal_dates_list:
        symbol_date_pairs.append(f"('{symbol}', '{date_str}')")
    in_clause_dates = ", ".join(symbol_date_pairs)

    query_signal_close = f"""
    SELECT symbol, trade_date, "close" 
    FROM stock_daily_kline 
    WHERE (symbol, trade_date) IN ({in_clause_dates}) 
    AND "close" IS NOT NULL;
    """
    df_signal_close = pd.read_sql(query_signal_close, engine)
    engine.dispose()

    # 构建信号日价格映射
    close_on_signal_map = {}
    for _, row in df_signal_close.iterrows():
        symbol = row['symbol']
        close_val = row['close']
        # 反向查找 code
        for code, info in stock_info_map.items():
            if info['symbol'] == symbol:
                close_on_signal_map[code] = close_val
                break
    print(f"✅ 成功获取 {len(close_on_signal_map)} 只股票的信号日收盘价")

    # --- 4.5 获取最新收盘价 ---
    engine = create_engine(DB_URI)
    latest_date_query = f"""
    WITH latest_trades AS (
        SELECT symbol, MAX(trade_date) AS latest_date
        FROM stock_daily_kline
        WHERE symbol IN ('{in_clause}')
        AND trade_date >= '{one_month_ago}'
        AND "close" IS NOT NULL
        GROUP BY symbol
    )
    SELECT lt.symbol, lt.latest_date, sdk."close" AS latest_close
    FROM latest_trades lt
    JOIN stock_daily_kline sdk ON lt.symbol = sdk.symbol AND lt.latest_date = sdk.trade_date;
    """
    df_latest_close = pd.read_sql(latest_date_query, engine)
    engine.dispose()

    # 构建最新价格映射
    latest_close_map = {}
    for _, row in df_latest_close.iterrows():
        symbol = row['symbol']
        latest_close = row['latest_close']
        latest_close_map[symbol] = latest_close

    # --- 4.6 最终筛选与涨幅计算 ---
    final_effective_stock_codes = []
    filtered_out = []

    for code in effective_stock_codes:
        symbol = stock_info_map[code]['symbol']
        if code not in close_on_signal_map:
            filtered_out.append(f"{stock_info_map[code]['name']} ({symbol}) - 无信号日价格")
            continue
        if symbol not in latest_close_map:
            filtered_out.append(f"{stock_info_map[code]['name']} ({symbol}) - 无最新价格")
            continue

        signal_close = close_on_signal_map[code]
        latest_close = latest_close_map[symbol]

        if latest_close > signal_close:
            final_effective_stock_codes.append(code)
        else:
            filtered_out.append(
                f"{stock_info_map[code]['name']} ({symbol}) - 信号日 {signal_close}, 最新价 {latest_close} ❌ 未上涨")

    print(f"✅ 最终通过趋势验证的有效股票：{len(final_effective_stock_codes)} 只")
    for reason in filtered_out:
        print(f" - {reason}")

    if len(final_effective_stock_codes) == 0:
        print("⚠️ 没有任何股票满足‘KDJ信号后股价上涨’的条件。")
        exit()

    # --- 4.7 计算涨幅百分比 ---
    gain_percentage_map = {}
    for code in final_effective_stock_codes:
        symbol = stock_info_map[code]['symbol']
        signal_close = close_on_signal_map[code]
        latest_close = latest_close_map[symbol]
        gain_pct = ((latest_close - signal_close) / signal_close) * 100
        gain_percentage_map[code] = round(gain_pct, 2)

    print(f"📈 已计算 {len(gain_percentage_map)} 只股票的涨幅百分比")

    # --- 4.8 获取30天交易数据 ---
    engine = create_engine(DB_URI)
    prefixed_stock_symbols = [stock_info_map[code]['symbol'] for code in final_effective_stock_codes]
    in_clause = "', '".join(prefixed_stock_symbols)

    query_kline = f"""
    SELECT symbol, trade_date, "close" 
    FROM stock_daily_kline 
    WHERE symbol IN ('{in_clause}') 
    AND trade_date >= '{one_month_ago}' 
    AND "close" IS NOT NULL 
    ORDER BY symbol, trade_date;
    """
    df_kline = pd.read_sql(query_kline, engine)
    engine.dispose()

    if df_kline.empty:
        print("⚠️ 未找到有效股票的交易数据（近30天）。")
        exit()

    # 格式化日期
    df_kline['trade_date'] = pd.to_datetime(df_kline['trade_date']).dt.strftime('%Y-%m-%d')
    trade_dates = sorted(df_kline['trade_date'].unique())
    print(f"✅ 共获取 {len(trade_dates)} 个交易日，覆盖范围：{trade_dates[0]} 至 {trade_dates[-1]}")

    # 构建价格映射表
    close_map = {}
    for _, row in df_kline.iterrows():
        symbol = row['symbol']
        date_str = row['trade_date']
        close_map.setdefault(symbol, {})[date_str] = row['close']

    # --- 4.9 获取KDJ信号高亮点 ---
    engine = create_engine(DB_URI)
    query_signals = f"""
    SELECT stock_code, archive_date::date AS archive_date 
    FROM app_stock_strategy_report 
    WHERE kdj_signal IS NOT NULL 
    AND archive_date >= '{one_month_ago}' 
    ORDER BY stock_code, archive_date;
    """
    df_signals = pd.read_sql(query_signals, engine)
    engine.dispose()

    highlight_map = {}
    for _, row in df_signals.iterrows():
        stock_code = row['stock_code']
        date_str = row['archive_date'].strftime('%Y-%m-%d')
        prefixed_symbol = add_exchange_prefix(stock_code)
        if prefixed_symbol in prefixed_stock_symbols:
            highlight_map[(prefixed_symbol, date_str)] = True
    print(f"✅ 共 {len(highlight_map)} 个 KDJ 信号点可用于高亮")

    # --- 4.10 获取MACD信号高亮点 ---
    query_macd_signals = f"""
    SELECT stock_code, archive_date::date AS archive_date, macd_12269_signal 
    FROM app_stock_strategy_report 
    WHERE macd_12269_signal IS NOT NULL 
    AND archive_date >= '{one_month_ago}' 
    ORDER BY stock_code, archive_date;
    """
    df_macd_signals = pd.read_sql(query_macd_signals, engine)
    engine.dispose()

    macd_highlight_map = {}
    for _, row in df_macd_signals.iterrows():
        stock_code = row['stock_code']
        date_str = row['archive_date'].strftime('%Y-%m-%d')
        signal_value = str(row['macd_12269_signal']).strip()
        prefixed_symbol = add_exchange_prefix(stock_code)

        if prefixed_symbol not in prefixed_stock_symbols:
            continue

        # 颜色映射逻辑
        if signal_value in ['下金叉', '下叉', '金叉', 'buy', '1', 'BUY', '正金叉']:
            color = "ADD8E6"  # 浅蓝色
            macd_highlight_map[(prefixed_symbol, date_str)] = color
        elif signal_value in ['上金叉', '上叉', '死叉', 'sell', '-1', 'SELL', '负金叉']:
            color = "9370DB"  # 紫色
            macd_highlight_map[(prefixed_symbol, date_str)] = color

    print(f"✅ 共 {len(macd_highlight_map)} 个 MACD 信号点可用于高亮")

    # --- 5. Excel 报告生成 (核心修改区) ---
    print("📝 正在生成 Excel 报告...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Close & Signal Report"

    # --- 5.1 标题行 (Row 1) ---
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1,
                         value="股票收盘价与多因子信号聚焦报告（近30天，仅展示有信号且价格持续上涨的股票）")
    title_cell.font = Font(bold=True, size=16, color="2E5488")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(trade_dates) + 3)  # 合并单元格列数+1
    ws.row_dimensions[1].height = 35

    # --- 5.2 图例说明行 (Row 2-4) ---
    ws.insert_rows(2)
    note_cell = ws.cell(row=2, column=1, value="📌 筛选逻辑：仅展示‘有KDJ信号’且‘信号后股价上涨’的股票。\n"
                                               "🎨 高亮说明：\n"
                                               "🔵 蓝色：MACD 零轴下金叉（买入信号）\n"
                                               "🟣 紫色：MACD 零轴上金叉（买入信号）\n"
                                               "🔴 红色：KDJ 信号（买入/卖出）\n"
                                               "✅ 所有股票均满足：信号日后价格上涨，确保动能有效。")
    note_cell.font = Font(bold=True, color="2E5488", size=12)
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=len(trade_dates) + 3)  # 合并单元格列数+1
    ws.row_dimensions[2].height = 80

    # --- 5.3 表头行 (Row 5) ---
    # 修改点：列顺序调整
    ws.cell(row=5, column=1, value="Stock Code")  # 第1列：股票代码
    ws.cell(row=5, column=2, value="Stock Name")  # 第2列：股票名称
    ws.cell(row=5, column=3, value="Signal to Latest Gain (%)")  # 第3列：涨幅

    # 日期表头从第4列开始
    for col_idx, date_str in enumerate(trade_dates, 4):
        ws.cell(row=5, column=col_idx, value=date_str)

    # --- 5.4 数据行 (Row 6+) ---
    row_idx = 6
    for stock_code in final_effective_stock_codes:
        name = stock_info_map[stock_code]['name']
        symbol = stock_info_map[stock_code]['symbol']

        # 第1列：股票代码
        ws.cell(row=row_idx, column=1, value=stock_code)

        # 第2列：股票名称
        ws.cell(row=row_idx, column=2, value=name)

        # 第3列：涨幅百分比
        gain_pct = gain_percentage_map[stock_code]
        ws.cell(row=row_idx, column=3, value=f"{gain_pct:+.2f}%")

        # 第4列起：交易日收盘价
        for col_idx, date_str in enumerate(trade_dates, 4):
            close_val = close_map.get(symbol, {}).get(date_str, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=close_val)

            # 高亮逻辑
            if (symbol, date_str) in macd_highlight_map:
                color_hex = macd_highlight_map[(symbol, date_str)]
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                cell.fill = fill
            elif (symbol, date_str) in highlight_map:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.fill = fill

        row_idx += 1

    # --- 6. 格式优化 ---
    total_cols = len(trade_dates) + 3  # 总列数+1
    for col_idx in range(1, total_cols + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = min(max(max_length + 2, 8), 25)
        ws.column_dimensions[column].width = adjusted_width

    # --- 7. 保存文件 ---
    today_str = datetime.now().strftime('%Y%m%d')
    excel_file = os.path.join(REPORT_OUTPUT_DIR, f"KDJ报告_{today_str}.xlsx")
    wb.save(excel_file)
    print(f"🎉 Excel 文件已生成：{excel_file}")
